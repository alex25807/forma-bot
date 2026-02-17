"""Export user history to Excel (.xlsx)."""

import io
import logging
from datetime import datetime

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

logger = logging.getLogger(__name__)

HEADER_FILL = PatternFill(start_color="1a1a2e", end_color="1a1a2e", fill_type="solid")
HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
ALT_FILL = PatternFill(start_color="f0f4f8", end_color="f0f4f8", fill_type="solid")
THIN_BORDER = Border(
    left=Side(style="thin", color="cccccc"),
    right=Side(style="thin", color="cccccc"),
    top=Side(style="thin", color="cccccc"),
    bottom=Side(style="thin", color="cccccc"),
)


def _style_header(ws, columns: list[str]):
    for col_idx, title in enumerate(columns, 1):
        cell = ws.cell(row=1, column=col_idx, value=title)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center")
        cell.border = THIN_BORDER


def _auto_width(ws):
    for col in ws.columns:
        max_len = 0
        col_letter = col[0].column_letter
        for cell in col:
            if cell.value:
                max_len = max(max_len, len(str(cell.value)))
        ws.column_dimensions[col_letter].width = min(max_len + 4, 50)


def generate_history_xlsx(
    daily_history: list[dict],
    weight_history: list[dict],
    profile: dict | None = None,
    first_name: str = "",
) -> bytes:
    """Return .xlsx file as bytes."""
    wb = Workbook()

    # ── Sheet 1: Daily log ────────────────────────────────────────
    ws_daily = wb.active
    ws_daily.title = "Дневник"
    cols_daily = ["Дата", "Утро", "Итог дня", "Отклонение", "Еда", "Обзор GPT"]
    _style_header(ws_daily, cols_daily)

    for i, rec in enumerate(daily_history, 2):
        ws_daily.cell(row=i, column=1, value=rec.get("log_date", ""))
        ws_daily.cell(row=i, column=2, value=rec.get("morning_state", ""))
        ws_daily.cell(row=i, column=3, value=rec.get("evening_result", ""))
        ws_daily.cell(row=i, column=4, value=rec.get("deviation_reason", ""))
        ws_daily.cell(row=i, column=5, value=rec.get("food_text", ""))
        ws_daily.cell(row=i, column=6, value=rec.get("gpt_review", ""))
        for c in range(1, 7):
            cell = ws_daily.cell(row=i, column=c)
            cell.border = THIN_BORDER
            cell.alignment = Alignment(wrap_text=True, vertical="top")
            if i % 2 == 0:
                cell.fill = ALT_FILL

    _auto_width(ws_daily)

    # ── Sheet 2: Weight log ───────────────────────────────────────
    ws_weight = wb.create_sheet("Вес")
    cols_weight = ["Дата", "Вес (кг)"]
    _style_header(ws_weight, cols_weight)

    for i, rec in enumerate(weight_history, 2):
        dt_str = rec.get("logged_at", "")[:10]
        ws_weight.cell(row=i, column=1, value=dt_str)
        ws_weight.cell(row=i, column=2, value=rec.get("weight_kg", ""))
        for c in range(1, 3):
            cell = ws_weight.cell(row=i, column=c)
            cell.border = THIN_BORDER
            cell.alignment = Alignment(horizontal="center")
            if i % 2 == 0:
                cell.fill = ALT_FILL

    _auto_width(ws_weight)

    # ── Sheet 3: Profile summary ──────────────────────────────────
    if profile:
        ws_profile = wb.create_sheet("Профиль")
        _style_header(ws_profile, ["Параметр", "Значение"])
        fields = [
            ("Пол", profile.get("gender", "")),
            ("Рост (см)", profile.get("height_cm", "")),
            ("Вес (кг)", profile.get("weight_kg", "")),
            ("Возраст", profile.get("age", "")),
            ("Активность", profile.get("activity", "")),
            ("Цель", profile.get("target", "")),
            ("Ограничения", profile.get("restrictions", "")),
            ("Ккал", profile.get("calories", "")),
            ("Белки (г)", profile.get("protein_g", "")),
            ("Жиры (г)", profile.get("fat_g", "")),
            ("Углеводы (г)", profile.get("carbs_g", "")),
        ]
        for i, (label, val) in enumerate(fields, 2):
            ws_profile.cell(row=i, column=1, value=label).border = THIN_BORDER
            ws_profile.cell(row=i, column=2, value=str(val) if val else "").border = THIN_BORDER
        _auto_width(ws_profile)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()
